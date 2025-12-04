"""Debt payoff calculator with snowball and avalanche methods."""

import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# Maximum months to prevent infinite loops
MAX_PAYOFF_MONTHS = 1200  # 100 years


def calculate_debt_payoff(debts, method="snowball", extra_payment=0):
    """
    Calculate debt payoff schedule using snowball or avalanche method.
    
    Args:
        debts: List of debt dicts with 'name', 'balance', 'interest_rate', 'min_payment'
        method: Payoff method ('snowball' or 'avalanche')
        extra_payment: Additional monthly payment to apply (default 0)
    
    Returns:
        DataFrame with monthly payoff schedule
    
    Raises:
        ValueError: If debts list is empty or contains invalid values
    """
    if not debts:
        raise ValueError("At least one debt is required")
    
    for debt in debts:
        if debt["balance"] <= 0:
            raise ValueError(f"Debt '{debt['name']}' must have a positive balance")
        if debt["interest_rate"] < 0:
            raise ValueError(f"Debt '{debt['name']}' interest rate cannot be negative")
        if debt["min_payment"] <= 0:
            raise ValueError(f"Debt '{debt['name']}' must have a positive minimum payment")
        
        # Check if minimum payment covers at least the first month's interest
        monthly_interest = debt["balance"] * (debt["interest_rate"] / 100) / 12
        if debt["min_payment"] <= monthly_interest and extra_payment == 0:
            raise ValueError(
                f"Debt '{debt['name']}': minimum payment ${debt['min_payment']:.2f} "
                f"does not cover monthly interest ${monthly_interest:.2f}. "
                f"Add extra payment to pay off this debt."
            )
    
    if extra_payment < 0:
        raise ValueError("Extra payment cannot be negative")
    
    # Make a deep copy of debts to avoid modifying the original
    debts = [{**d} for d in debts]
    
    # Sort debts based on selected method
    if method == "snowball":
        debts = sorted(debts, key=lambda x: x["balance"])  # Smallest balance first
    elif method == "avalanche":
        debts = sorted(debts, key=lambda x: x["interest_rate"], reverse=True)  # Highest rate first

    total_interest_paid = 0
    results = []
    month = 1

    # Iterate until all debts are paid off (with max months protection)
    while any(debt["balance"] > 0 for debt in debts) and month <= MAX_PAYOFF_MONTHS:
        monthly_summary = {"Month": month, "Total Payment": 0, "Total Interest Paid": total_interest_paid}
        extra_remaining = extra_payment

        for debt in debts:
            if debt["balance"] <= 0:
                continue  # Skip paid-off debts

            interest = debt["balance"] * (debt["interest_rate"] / 100) / 12
            minimum_payment = debt["min_payment"]

            if debt["balance"] + interest <= minimum_payment:
                payment = debt["balance"] + interest
            else:
                payment = minimum_payment + (extra_remaining if extra_remaining > 0 else 0)

            extra_remaining -= max(0, payment - (interest + minimum_payment))
            principal_payment = payment - interest
            debt["balance"] -= principal_payment

            total_interest_paid += interest

            # Add debt details to summary
            monthly_summary[f"Debt {debt['name']} Balance"] = debt["balance"]
            monthly_summary[f"Debt {debt['name']} Payment"] = payment
            monthly_summary[f"Debt {debt['name']} Interest"] = interest

        monthly_summary["Total Payment"] = sum(monthly_summary[f"Debt {debt['name']} Payment"] for debt in debts if f"Debt {debt['name']} Payment" in monthly_summary)
        monthly_summary["Total Interest Paid"] = total_interest_paid

        results.append(monthly_summary)
        month += 1

    return pd.DataFrame(results)

def plot_debt_payoff(df, file_name):
    """Generate and save debt payoff progress chart.
    
    Args:
        df: DataFrame with payoff schedule
        file_name: Output file path for the chart image
    """
    # Plot total debt balance over time
    plt.figure(figsize=(12, 7))
    for col in df.columns:
        if "Balance" in col:
            plt.plot(df["Month"], df[col], label=col.replace("Debt ", "").replace(" Balance", ""))
    plt.title("Debt Payoff Progress")
    plt.xlabel("Month")
    plt.ylabel("Remaining Balance ($)")
    plt.legend(loc="upper right")
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(file_name)
    plt.close()

def auto_adjust_column_width(file_name):
    """Auto-adjust column widths in Excel file to fit content.
    
    Args:
        file_name: Path to the Excel file
    """
    workbook = load_workbook(file_name)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for col in sheet.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[col_letter].width = adjusted_width
    workbook.save(file_name)

def embed_chart_in_excel(file_name, image_file):
    """Embed chart image into Excel file.
    
    Args:
        file_name: Path to the Excel file
        image_file: Path to the chart image file
    """
    workbook = load_workbook(file_name)
    graph_sheet_name = "Graph"
    if graph_sheet_name not in workbook.sheetnames:
        workbook.create_sheet(graph_sheet_name)
    chart_sheet = workbook[graph_sheet_name]

    img = Image(image_file)
    img.anchor = "A1"
    chart_sheet.add_image(img)

    workbook.save(file_name)

def export_to_excel(df, file_name):
    """Export debt payoff schedule to Excel with formatting.
    
    Args:
        df: DataFrame with payoff schedule
        file_name: Output Excel file path
    """
    with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Debt Payoff Schedule")
    workbook = load_workbook(file_name)
    sheet = workbook["Debt Payoff Schedule"]

    # Apply dollar formatting to numeric columns
    for col_name in df.columns:
        if "Payment" in col_name or "Balance" in col_name or "Interest" in col_name or col_name in ["Total Payment", "Total Interest Paid"]:
            col_letter = sheet.cell(row=1, column=df.columns.get_loc(col_name) + 1).column_letter
            for row in range(2, sheet.max_row + 1):
                cell = sheet[f"{col_letter}{row}"]
                cell.number_format = '"$"#,##0.00'

    workbook.save(file_name)

if __name__ == "__main__":
    # Input debts
    debts = []
    print("Enter your debts (name, balance, interest rate, minimum payment). Type 'done' when finished.")
    while True:
        name = input("Debt name: ")
        if name.lower() == "done":
            break
        balance = float(input(f"Enter balance for {name}: "))
        interest_rate = float(input(f"Enter interest rate for {name} (in %): "))
        min_payment = float(input(f"Enter minimum payment for {name}: "))
        debts.append({"name": name, "balance": balance, "interest_rate": interest_rate, "min_payment": min_payment})

    # Input payoff method and extra payment
    method = input("Choose payoff method ('snowball' or 'avalanche'): ").lower()
    extra_payment = float(input("Enter extra monthly payment (optional, default is 0): ") or 0)
    file_name = input("Enter the base name for the output files (e.g., 'debt_payoff'): ")

    # Calculate debt payoff and export results
    df = calculate_debt_payoff(debts, method, extra_payment)
    image_file = f"{file_name}.png"
    excel_file = f"{file_name}.xlsx"

    plot_debt_payoff(df, image_file)
    export_to_excel(df, excel_file)
    embed_chart_in_excel(excel_file, image_file)
    auto_adjust_column_width(excel_file)

    print(f"Debt payoff schedule saved to {excel_file} with an amortization graph embedded.")
