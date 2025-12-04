"""Mortgage calculator with amortization schedule and Excel export."""

import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image


def mortgage_calculator(principal, interest_rate, loan_term, property_tax=0, insurance=0, pmi=0, extra_payment=0):
    """
    Calculate mortgage amortization schedule.
    
    Args:
        principal: Loan principal amount
        interest_rate: Annual interest rate (percentage)
        loan_term: Loan term in years
        property_tax: Monthly property tax (default 0)
        insurance: Monthly homeowner's insurance (default 0)
        pmi: Monthly PMI (default 0)
        extra_payment: Extra monthly payment toward principal (default 0)
    
    Returns:
        DataFrame with monthly amortization schedule
    
    Raises:
        ValueError: If principal, interest rate, or loan term is invalid
    """
    if principal <= 0:
        raise ValueError("Principal must be positive")
    if interest_rate < 0:
        raise ValueError("Interest rate cannot be negative")
    if loan_term <= 0:
        raise ValueError("Loan term must be positive")
    if property_tax < 0 or insurance < 0 or pmi < 0 or extra_payment < 0:
        raise ValueError("Property tax, insurance, PMI, and extra payment cannot be negative")
    
    # Monthly interest rate and total number of payments
    monthly_rate = (interest_rate / 100) / 12
    total_payments = loan_term * 12

    # Calculate base monthly payment (excluding taxes and insurance)
    # Handle 0% interest rate edge case
    if monthly_rate == 0:
        base_payment = principal / total_payments
    else:
        base_payment = principal * monthly_rate / (1 - (1 + monthly_rate) ** -total_payments)

    # Include property tax, insurance, and PMI
    monthly_payment = base_payment + property_tax + insurance + pmi

    # Amortization schedule
    balance = principal
    total_interest_paid = 0
    results = []

    for month in range(1, total_payments + 1):
        interest = balance * monthly_rate
        principal_paid = base_payment - interest
        total_interest_paid += interest
        balance -= (principal_paid + extra_payment)
        balance = max(balance, 0)  # Prevent negative balances

        results.append({
            "Month": month,
            "Monthly Payment": monthly_payment + extra_payment if balance > 0 else 0,
            "Principal Paid": principal_paid + extra_payment if balance > 0 else 0,
            "Interest Paid": interest if balance > 0 else 0,
            "Property Tax": property_tax,
            "Insurance": insurance,
            "PMI": pmi,
            "Total Interest Paid": total_interest_paid,
            "Remaining Balance": balance
        })

        # Stop if the loan is paid off early
        if balance <= 0:
            break

    # Convert to DataFrame
    df = pd.DataFrame(results)
    return df

def plot_mortgage_amortization(df, file_name):
    """Generate and save mortgage amortization chart.
    
    Args:
        df: DataFrame with amortization schedule
        file_name: Output file path for the chart image
    """
    # Plot principal vs. interest breakdown
    plt.figure(figsize=(12, 7))
    plt.plot(df["Month"], df["Remaining Balance"], label="Remaining Balance", color="blue")
    plt.plot(df["Month"], df["Total Interest Paid"], label="Total Interest Paid", linestyle="--", color="orange")
    plt.title("Mortgage Amortization Over Time")
    plt.xlabel("Month")
    plt.ylabel("Amount ($)")
    plt.legend(loc="upper right")
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(file_name)
    plt.close()

def auto_adjust_column_width(file_name):
    """Auto-adjust column widths in Excel file to fit content.
    
    Args:
        file_name: Path to the Excel file
    """
    workbook = load_workbook(file_name)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for col in sheet.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[col_letter].width = adjusted_width
    workbook.save(file_name)

def embed_chart_in_excel(file_name, image_file):
    """Embed chart image into Excel file.
    
    Args:
        file_name: Path to the Excel file
        image_file: Path to the chart image file
    """
    workbook = load_workbook(file_name)
    graph_sheet_name = "Graph"
    if graph_sheet_name not in workbook.sheetnames:
        workbook.create_sheet(graph_sheet_name)
    chart_sheet = workbook[graph_sheet_name]

    img = Image(image_file)
    img.anchor = "A1"
    chart_sheet.add_image(img)

    workbook.save(file_name)

def export_to_excel(df, file_name):
    """Export amortization schedule to Excel with formatting.
    
    Args:
        df: DataFrame with amortization schedule
        file_name: Output Excel file path
    """
    with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Amortization Schedule")
    workbook = load_workbook(file_name)
    sheet = workbook["Amortization Schedule"]

    dollar_columns = ["Monthly Payment", "Principal Paid", "Interest Paid", "Property Tax", "Insurance", "PMI", "Total Interest Paid", "Remaining Balance"]
    for col_name in dollar_columns:
        if col_name in df.columns:
            col_letter = sheet.cell(row=1, column=df.columns.get_loc(col_name) + 1).column_letter
            for row in range(2, sheet.max_row + 1):
                cell = sheet[f"{col_letter}{row}"]
                cell.number_format = '"$"#,##0.00'

    workbook.save(file_name)

if __name__ == "__main__":
    while True:
        try:
            principal = float(input("Enter the loan principal amount: "))
            interest_rate = float(input("Enter the annual interest rate (in %): "))
            loan_term = int(input("Enter the loan term (in years): "))
            property_tax = float(input("Enter the monthly property tax (optional, default is 0): ") or 0)
            insurance = float(input("Enter the monthly homeowner’s insurance (optional, default is 0): ") or 0)
            pmi = float(input("Enter the monthly PMI (Private Mortgage Insurance, optional, default is 0): ") or 0)
            extra_payment = float(input("Enter the extra monthly payment (optional, default is 0): ") or 0)
            file_name = input("Enter the base name for the output files (e.g., 'mortgage'): ")
            break
        except Exception as e:
            print(f"Error: {e}. Please try again.")

    image_file = f"{file_name}.png"
    excel_file = f"{file_name}.xlsx"

    df = mortgage_calculator(principal, interest_rate, loan_term, property_tax, insurance, pmi, extra_payment)
    plot_mortgage_amortization(df, image_file)
    export_to_excel(df, excel_file)
    embed_chart_in_excel(excel_file, image_file)
    auto_adjust_column_width(excel_file)

    print(f"Mortgage details saved to {excel_file} with an amortization graph embedded.")
